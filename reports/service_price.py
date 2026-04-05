"""Service: Price list export (XLSX).

This module provides a service page that allows staff to export a price list grouped by brand.

Export rules:
- Group by brand, then sort by SKU.
- Options:
  - all products vs only in-stock
  - show qty vs hide qty
  - brand filter
- XLSX includes a product image (embedded).

All code comments are in English per project conventions.
"""

from __future__ import annotations

import io
import re
from pathlib import Path
from collections import defaultdict
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional

import requests
from django.contrib.auth.decorators import user_passes_test
from django.http import HttpRequest, HttpResponse
from django.shortcuts import render
from django.utils import timezone

from b2b.models import Brand, Product


def _is_staff(user) -> bool:
    return bool(getattr(user, "is_authenticated", False) and getattr(user, "is_staff", False))


def _parse_int_list(values: Iterable[str]) -> List[int]:
    out: List[int] = []
    for v in values:
        try:
            out.append(int(v))
        except Exception:
            continue
    return out


def _build_absolute_url(request: HttpRequest, url: str) -> str:
    url = (url or "").strip()
    if not url:
        return ""
    if url.startswith("http://") or url.startswith("https://"):
        return url
    return request.build_absolute_uri(url)


@dataclass(frozen=True)
class PriceRow:
    brand_name: str
    sku: str
    name: str
    wholesale_price: float
    stock_qty: int
    image_url: str


@user_passes_test(_is_staff)
def service_price(request: HttpRequest) -> HttpResponse:
    """Settings page for exporting a price list."""

    scope = (request.GET.get("scope") or "in_stock").strip()  # in_stock | all
    show_qty = (request.GET.get("show_qty") or "0").strip()  # 1|0
    brand_ids = _parse_int_list(request.GET.getlist("brand_ids"))

    brands = Brand.objects.order_by("name")

    # Preview counts.
    qs = Product.objects.select_related("brand").filter(is_active=True, wholesale_price__gt=0)
    if scope == "in_stock":
        qs = qs.filter(stock_qty__gt=0)
    if brand_ids:
        qs = qs.filter(brand_id__in=brand_ids)

    total_count = qs.count()
    in_stock_count = qs.filter(stock_qty__gt=0).count()

    return render(
        request,
        "reports/service_price.html",
        {
            "scope": scope,
            "show_qty": show_qty,
            "brands": brands,
            "selected_brand_ids": set(brand_ids),
            "total_count": total_count,
            "in_stock_count": in_stock_count,
        },
    )


@user_passes_test(_is_staff)
def service_price_export(request: HttpRequest) -> HttpResponse:
    """Generate and return the XLSX price list."""

    data = request.POST if request.method == "POST" else request.GET

    scope = (data.get("scope") or "in_stock").strip()  # in_stock | all
    show_qty = (data.get("show_qty") or "0").strip()  # 1|0
    brand_ids = _parse_int_list(data.getlist("brand_ids"))

    qs = Product.objects.select_related("brand").filter(is_active=True, wholesale_price__gt=0)
    if scope == "in_stock":
        qs = qs.filter(stock_qty__gt=0)
    if brand_ids:
        qs = qs.filter(brand_id__in=brand_ids)

    rows: List[PriceRow] = []
    for p in qs.order_by("brand__name", "sku"):
        brand_name = p.brand.name if p.brand else "Без бренду"
        rows.append(
            PriceRow(
                brand_name=brand_name,
                sku=(p.sku or "").strip(),
                name=(getattr(p, "name_with_weight", None) or p.name or "").strip(),
                wholesale_price=float(p.wholesale_price or 0),
                stock_qty=int(p.stock_qty or 0),
                image_url=_build_absolute_url(request, getattr(p, "main_image_url", "") or ""),
            )
        )

    xlsx_bytes = _build_price_xlsx(rows=rows, show_qty=(show_qty == "1"))

    ts = timezone.localtime(timezone.now()).strftime("%Y-%m-%d_%H%M")
    filename = f"price_{ts}.xlsx"

    resp = HttpResponse(
        xlsx_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _build_price_xlsx(*, rows: List[PriceRow], show_qty: bool) -> bytes:
    """Build XLSX content as bytes."""

    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from PIL import Image as PILImage

    def pil_to_png_bytes(img: PILImage.Image, *, max_size: tuple[int, int]) -> bytes:
        """Convert a Pillow image into PNG bytes suitable for XLSX embedding.

        We embed higher-resolution PNG bytes (e.g. 320x320) and then scale the image down
        via openpyxl's width/height. This keeps the result sharp in Excel.
        """
        # Excel can be picky with alpha; flatten to white.
        if img.mode in ("RGBA", "LA"):
            bg = PILImage.new("RGB", img.size, (255, 255, 255))
            bg.paste(img, mask=img.split()[-1])
            img = bg
        else:
            img = img.convert("RGB")

        # High-quality downscale
        resampling = getattr(PILImage, "Resampling", PILImage)
        resample = getattr(resampling, "LANCZOS", getattr(PILImage, "LANCZOS", None))
        img = img.copy()
        try:
            img.thumbnail(max_size, resample=resample)
        except TypeError:
            # Pillow < 9 compatibility
            img.thumbnail(max_size)

        buf = io.BytesIO()
        # PNG is lossless; optimize reduces size without reducing quality
        img.save(buf, format="PNG", optimize=True)
        return buf.getvalue()

    image_cache: Dict[str, Optional[bytes]] = {}

    def fetch_image_png(url: str, *, max_size: tuple[int, int]) -> Optional[bytes]:
        if not url:
            return None

        cache_key = f"{url}|{max_size[0]}x{max_size[1]}"
        if cache_key in image_cache:
            return image_cache[cache_key]

        # WordPress often stores thumbnails as "...-300x300.jpg". Try original first.
        urls_to_try = [url]
        try:
            m = re.search(r"-(\d+)x(\d+)(\.(?:png|jpe?g|webp))(\?.*)?$", url, flags=re.IGNORECASE)
            if m:
                hi = url[: m.start()] + m.group(3) + (m.group(4) or "")
                if hi != url:
                    urls_to_try = [hi, url]
        except Exception:
            urls_to_try = [url]

        for u in urls_to_try:
            try:
                resp = requests.get(u, timeout=15)
                resp.raise_for_status()
                pil = PILImage.open(io.BytesIO(resp.content))
                png = pil_to_png_bytes(pil, max_size=max_size)
                image_cache[cache_key] = png
                return png
            except Exception:
                continue

        image_cache[cache_key] = None
        return None
        cache_key = f"{url}|{max_size[0]}x{max_size[1]}"
        if cache_key in image_cache:
            return image_cache[cache_key]
        try:
            resp = requests.get(url, timeout=15)
            resp.raise_for_status()
            pil = PILImage.open(io.BytesIO(resp.content))
            png = pil_to_png_bytes(pil, max_size=max_size)
            image_cache[cache_key] = png
            return png
        except Exception:
            image_cache[cache_key] = None
            return None
        if url in image_cache:
            return image_cache[url]
        try:
            resp = requests.get(url, timeout=10)
            resp.raise_for_status()
            pil = PILImage.open(io.BytesIO(resp.content))
            png = pil_to_png_bytes(pil, max_size=(1200, 300))
            image_cache[url] = png
            return png
        except Exception:
            image_cache[url] = None
            return None

    def placeholder_png() -> bytes:
        pil = PILImage.new("RGB", (320, 320), (240, 240, 240))
        return pil_to_png_bytes(pil, max_size=(320, 320))

    wb = Workbook()
    ws = wb.active
    ws.title = "Прайс"

    # Columns
    col_headers = ["Фото", "SKU", "Назва", "Ціна, грн"]
    if show_qty:
        col_headers.append("К-сть")

    widths = [12, 14, 58, 14] + ([8] if show_qty else [])
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Styles
    title_font = Font(bold=True, size=14)
    group_font = Font(bold=True, size=12, color="FFFFFF")
    header_font = Font(bold=True)

    group_fill = PatternFill("solid", fgColor="1F2937")
    header_fill = PatternFill("solid", fgColor="E5E7EB")

    wrap = Alignment(wrap_text=True, vertical="top")
    middle = Alignment(vertical="center")

    def load_logo_png() -> Optional[bytes]:
        """Load a local logo image as PNG bytes (best-effort)."""
        # Try to locate a PNG/JPG logo inside Django staticfiles or project tree.
        candidates = [
            "b2b/logo.png",
            "b2b/img/logo.png",
            "b2b/images/logo.png",
            "img/logo.png",
            "images/logo.png",
            "logo.png",
            "static/b2b/logo.png",
            "static/b2b/img/logo.png",
            "static/img/logo.png",
            "static/logo.png",
        ]
        path = None
        try:
            from django.contrib.staticfiles import finders  # type: ignore

            for rel in candidates:
                found = finders.find(rel)
                if found:
                    path = found
                    break
        except Exception:
            path = None

        if not path:
            # Fallback: search in BASE_DIR recursively (shallow).
            try:
                from django.conf import settings as dj_settings  # type: ignore
                base = Path(dj_settings.BASE_DIR)
                for rel in candidates:
                    cand = base / rel
                    if cand.exists():
                        path = str(cand)
                        break
                if not path:
                    for patt in ["*logo*.png", "*logo*.jpg", "*logo*.jpeg"]:
                        hits = list(base.glob(patt))
                        if hits:
                            path = str(hits[0])
                            break
            except Exception:
                path = None

        if not path:
            return None

        try:
            pil = PILImage.open(path)
            png = pil_to_png_bytes(pil, max_size=(1200, 300))
            return png
        except Exception:
            return None

    # Header (logo + call-to-action link)
    ncols = len(col_headers)

    # Row 1: logo (centered approximately)
    logo_png = load_logo_png()
    if logo_png:
        bio_logo = io.BytesIO(logo_png)
        bio_logo.name = "logo.png"
        logo_xl = XLImage(bio_logo)
        # Make the logo slightly larger than product thumbnails.
        logo_xl.width = 220
        logo_xl.height = 80
        anchor_col = min(ncols, max(1, (ncols + 1) // 2) + 1)
        logo_xl.anchor = f"{get_column_letter(anchor_col)}1"
        ws.add_image(logo_xl)
    ws.row_dimensions[1].height = 60

    # Row 2: title
    ws["A2"].value = "Прайс"
    ws["A2"].font = title_font
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.row_dimensions[2].height = 22

    # Row 3: CTA text
    ws["A3"].value = "Реєструйтесь на нашому B2B-порталі"
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    ws.row_dimensions[3].height = 18

    # Row 4: hyperlink
    link_cell = ws["A4"]
    link_cell.value = "b2b.herabuna.com.ua"
    link_cell.hyperlink = "https://b2b.herabuna.com.ua"
    link_cell.font = Font(color="0563C1", underline="single")
    link_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)
    ws.row_dimensions[4].height = 18

    # Blank spacer row
    ws.row_dimensions[5].height = 8

    # Group rows by brand
    by_brand: Dict[str, List[PriceRow]] = defaultdict(list)
    for r in rows:
        by_brand[r.brand_name].append(r)

    current_row = 6

    for brand_name in sorted(by_brand.keys(), key=lambda s: s.lower()):
        items = sorted(by_brand[brand_name], key=lambda x: (x.sku or "", x.name))

        # Brand header
        ws.cell(row=current_row, column=1, value=brand_name).font = group_font
        ws.cell(row=current_row, column=1).fill = group_fill
        ws.cell(row=current_row, column=1).alignment = middle
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(col_headers))
        ws.row_dimensions[current_row].height = 20
        for c in range(1, len(col_headers) + 1):
            ws.cell(row=current_row, column=c).fill = group_fill

        current_row += 1

        # Column header within group
        for c, h in enumerate(col_headers, start=1):
            cell = ws.cell(row=current_row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = middle
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Items
        for it in items:
            ws.row_dimensions[current_row].height = 54

            ws.cell(row=current_row, column=2, value=it.sku).alignment = middle
            name_cell = ws.cell(row=current_row, column=3, value=it.name)
            name_cell.alignment = wrap

            price_cell = ws.cell(row=current_row, column=4, value=float(it.wholesale_price))
            price_cell.number_format = "0.00"
            price_cell.alignment = middle

            if show_qty:
                qty_cell = ws.cell(row=current_row, column=5, value=int(it.stock_qty))
                qty_cell.alignment = middle

            # Image embedding: openpyxl requires a file-like object with a 'name' extension.
            png = fetch_image_png(it.image_url, max_size=(320, 320)) or placeholder_png()
            bio = io.BytesIO(png)
            bio.name = "image.png"
            xl_img = XLImage(bio)
            xl_img.width = 72
            xl_img.height = 72
            xl_img.anchor = f"A{current_row}"
            ws.add_image(xl_img)

            current_row += 1

        current_row += 1  # blank row between brands

    ws.freeze_panes = "A6"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
